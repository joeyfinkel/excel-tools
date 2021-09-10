import collections
import pandas as pd
import pprint
import itertools
from openpyxl import Workbook, load_workbook
from Templates.item_template import ItemTemplate
import Templates.ItemTemplate.headers as headers
import Templates.ItemTemplate.columns as columns

item = ItemTemplate()
excel = Workbook()
new_wb = Workbook()
header = headers.Headers()

excel_sheet = excel.active
new_sheet = new_wb.active

csv_file = (
    r"C:\Users\joeyf\Desktop\Work\For Excel Automation\americanbuyergroupasindata.csv"
)
new_file = r"C:\Users\joeyf\Desktop\Work\For Excel Automation\ABG ASIN Data.xlsx"

# original_book = item.csv_to_xlsx(csv_file, new_file)
original_book = load_workbook(r"C:\Users\joeyf\Desktop\Test.xlsx")
sheet = original_book.active

ref_headers = item.ref_headers()

column = [
    item.get_column(sheet, idx, 2)
    for idx, el in enumerate(item.headers_from_sheet(sheet))
]
for idx, el in enumerate(item.headers_from_sheet(sheet)):
    column[idx].insert(0, el)

all_column = list(itertools.chain.from_iterable(column))
print(column)
result = [all_column[idx :: sheet.max_row] for idx, el in enumerate(all_column)]
result = [r for r in result if len(r) == len(result[0])]

to_keep = ["ID", "Item"]
final = []
for word in result[0]:
    if word in to_keep:
        final.append(word)
    else:
        final.append("")

result[0] = final
headers = result[0]

insert_empty = headers.index("")

for r in result:
    r[insert_empty] = ""

result = [r for r in result if result.index(r) > 0]
result = [list(item.insert_every_n(headers, r, 1)) for r in result]
# print(result)

# Creates a row ready to add to sheet
data = []
for res in result:
    row2 = [res[res.index(x) + 1] for x in headers]
    data.append(list(item.insert_every_n(headers, row2, 1)))

# data = [print([i for i in d]) for d in data]

[print([c for c in column[1::2] if c != ""]) for column in data]

# header.populate(new_sheet, headers)
# columns.Columns().populate(new_sheet, data)
# new_wb.save(r'C:\Users\joeyf\Desktop\Test2.xlsx')
