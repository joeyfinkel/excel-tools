import collections
import pprint
from Base import Base
from itertools import takewhile
from ItemTemplate import ItemTemplate
from openpyxl import Workbook, load_workbook

item = ItemTemplate()
new_wb = Workbook()
# original_book = load_workbook('C:/Users/joeyf/Desktop/Work/For Excel Automation/ciataasindata (1).xlsx')
original_book = load_workbook('Grades.xlsx')

sheet = original_book.active
new_sheet = new_wb.active

ref_headers = ['ID', 'Name', 'English', 'Math']

column = [item.get_column(sheet, idx, 2) for idx, el in enumerate(ref_headers)]
id_row = [s.index('ID') for s in sheet.iter_rows(max_row = 1, values_only = True)]
headers = [item.create_heading_list(el, len(column[idx])) for idx, el in enumerate(ref_headers)]
column_header_list = [list(item.insert_every_n(headers[idx], column[idx], k = 1)) for idx, el in enumerate(ref_headers)]
attributes = [item.set_attributes(column_header_list[idx]) for idx, el in enumerate(column_header_list)]
res = [el[idx] for idx, el in enumerate(attributes)]

print(attributes[0:2][0:2])
# print(attributes[0][1])
# print(attributes[0][2])
print()
new_res = []

# for idx, el in enumerate(attributes):
#     # counter = len(attributes[idx])
#     # new_res.append([attributes[idx][c] for c in range(counter)])
#     # # print(attributes[idx][idx+1])
#     for i, e in enumerate(attributes[idx]):
#         print(i, attributes[idx][i])
# print(new_res)
# print()
# for idx, el in enumerate(attributes):
#     print(el[idx])
final = {}
# for ID in id_row:
#     for r in res:
#         for k, v in r.items():
#             final.setdefault(ID, {}).update({k:v})
# print(final)