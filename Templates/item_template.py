import os
import csv
import eel
import pandas as pd
from itertools import chain
from openpyxl import Workbook, load_workbook
from .base import Base

@eel.expose
class ItemTemplate(Base):
    PACKAGE = 'package'
    PRODUCT = 'item'
    HEADERS = {
        'general_headers': ['asin', 'manufacturer', 'title', 'brand', 'color', 'size', 'description', 'upcList'],
        'dimensions': ['height', 'length', 'width', 'weight']
    }

    def headers_from_sheet(self, sheet) -> list:
        ''' Creates a list of the original headers '''
        headers = []

        for col in sheet.iter_rows(max_row = 1, values_only = True):
            [headers.append(c) for c in col]

        return headers

    # headers
    def create_dim_list(self, specific_type, dims) -> list:
        '''
        Creates dimension list
        :param specific_type: product or package
        :param dims: list of dimensions
        :return product/package list
        '''
        package_dims = []

        [package_dims.append(f'{specific_type}{i.capitalize()}') for i in dims]

        return package_dims

    def create_feature_list(self) -> list:
        features = []

        [features.append('feature{}'.format(i)) for i in range(7)]

        return features

    def ref_headers(self) -> list:
        ''' return: list of all headers to find '''
        dims = self.HEADERS['dimensions']
        package_dims = self.create_dim_list(self.PACKAGE, dims)
        product_dims = self.create_dim_list(self.PRODUCT, dims)
        features = self.create_feature_list()

        return list(chain(self.HEADERS['general_headers'], package_dims, product_dims, features))

    @eel.expose
    def csv_to_xlsx(self, csv_filename, new_name):
        wb = Workbook()
        sheet = wb.active

        with open(csv_filename) as f:
            reader = csv.reader(f, delimiter=',')
            [sheet.append(row) for row in reader]

        wb.save(new_name)

        return wb

    def remove_file(self, filename):
        if os.path.exists(filename):
            os.remove(filename)
        else:
            print(f'{filename} not found')
