import itertools
import Templates.utils as utils
import Templates.item_template as item_template
from openpyxl import worksheet


class Columns:
    template = item_template.ItemTemplate()

    def get_columns_without_headers(self, sheet: worksheet, column) -> list:
        """Gets columns from sheet without the headers

        :param sheet: Sheet to get columns from
        :type sheet: worksheet
        :param column: Column to get data from
        :type column: int
        :return: 2 dimensional list of all the columns
        :rtype: list
        """
        col = []

        for i in sheet.iter_rows(min_row=2, values_only=True):
            col.append(i[column])

        return col

    def get_columns_with_header(self, sheet: worksheet, column) -> list:
        """Gets columns from sheet with the headers

        :param sheet: Sheet to get columns from
        :type sheet: worksheet
        :param column: Column to get data from
        :type column: int
        :return: 2 dimensional list of all the columns
        :rtype: list
        """
        col = []

        for i in sheet.iter_rows(min_row=1, values_only=True):
            col.append(i[column])

        return col

    def column_list_breakdown(self, sheet: worksheet, column: list) -> list:
        result = [column[idx :: sheet.max_row] for idx, _ in enumerate(column)]

        return [res for res in result if len(res) == len(result[0])]

    def get_columns_by_row(self, sheet: worksheet, file_data: list) -> list:
        """Gets column information row by row. Each row is put in it's own list.

        :param sheet: Sheet to get headers from
        :type sheet: worksheet
        :param file_data: List of all data from sheet
        :type file_data: list
        :return: List of all the headers from the sheet
        :rtype: list
        """
        column = list(itertools.chain.from_iterable(file_data))

        return self.column_list_breakdown(sheet, column)

    def get_total_columns(self, file: str) -> int:
        """Gets total number of columns in the sheet

        :param file: Path of the file
        :type file: str
        :return: Amount of columns in the sheet
        :rtype: int
        """
        return utils.Utils().init_worksheet(file).max_column

    def insert_empty_string(self, data: list, headers_list: list) -> list:
        """Inserts an empty to string in a list at specified index

        :param data: List to manipulate
        :type data: list
        :param index: Index of list to insert string
        :type index: int
        :return: List with empty strings at specified index
        :rtype: list
        """

        # Only do if headers is less than headers from sheet
        data[0] = headers_list
        headers = data[0]
        insert_index = headers.index("")

        for d in data:
            d[insert_index] = ""

        result = [d for d in data if data.index(d) > 0]
        result = [list(self.template.insert_every_n(headers, r, 1)) for r in result]

        return result

    def create_final_data(self, data_list: list, headers: list) -> list:
        """Creates a list of columns to add to the new sheet based on the headers the user clicked

        :param file: Path of the file
        :type file: str
        :param sheet: Sheet to get columns from
        :type sheet: worksheet
        :param data_list: List of headers user clicked on
        :type data_list: list
        :return: New list of columns based on final_headers
        :rtype: list
        """
        result = []

        for data in data_list:
            row = [data[data.index(x) + 1] for x in headers]
            result.append(list(self.template.insert_every_n(headers, row, 1)))

        return result

    def populate(self, sheet: worksheet, columns: list) -> None:
        """Populates the columns on the new sheet

        :param sheet: Sheet to add headers to
        :type sheet: worksheet
        :param columns: List of columns from original sheet
        :type columns: list
        """
        [sheet.append([c for c in column[1::2] if c != ""]) for column in columns]
