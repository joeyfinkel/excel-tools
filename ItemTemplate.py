from itertools import chain
from Base import Base
from openpyxl import Workbook, load_workbook

class ItemTemplate(Base):
    PACKAGE = 'package'
    PRODUCT = 'product'
    HEADERS = {
        'general_headers': ['asin', 'manufacturer', 'title', 'brand', 'color', 'size', 'description', 'upcList'],
        'dimensions': ['height', 'length', 'width', 'weight']
    }

    def headers_from_sheet(self, sheet) -> list:
        '''
        creates a list of the original headers
        '''
        headers = []

        for col in sheet.iter_rows(max_row = 1, values_only = True):
            [headers.append(c) for c in col]

        return headers

    def check_headers(self, sheet, headers, new_headers) -> list:
        '''
        checks if new specified headers are in the original workbook
        :param sheet: active sheet
        :param header: list of headers
        :return columns values of headers
        '''
        columns = []

        for header in headers:
            for new in new_headers:
                if header == new:
                    columns.append(self.get_column(sheet, headers.index(header), row = 1))

        return columns

    # headers
    def create_dim_list(self, specific_type, dims) -> list:
        '''
        creates dimension list
        :param specific_type: product or package
        :param dims: list of dimensions
        :return product/package list
        '''
        package_dims = []

        [package_dims.append('{}{}' .format(specific_type, i.capitalize())) for i in dims]

        return package_dims

    def create_feature_list(self) -> list:
        features = []

        [features.append('feature{}'.format(i)) for i in range(7)]

        return features

    def ref_headers(self) -> list:
        '''
        return: list of all headers to find
        '''
        dims = self.HEADERS['dimensions']
        package_dims = self.create_dim_list(self.PACKAGE, dims)
        product_dims = self.create_dim_list(self.PRODUCT, dims)
        features = self.create_feature_list()

        return list(chain(self.HEADERS['general_headers'], package_dims, product_dims, features))

    def compare_headers(self, lst1, lst2) -> list:
        column = []

        for l1 in range(len(lst1)):
            for l2 in range(len(lst2)):
                if lst1[l1] in lst2[l2]:
                    column.append(lst1[l1])
                    break

        return column

    def search_headers(self, sheet, header):
        # go through ref headers and headers from sheet
        # when the values equal each other get the column values of that header and add it to a dict
        # where the header is the key and value would be an array all the column values
        headers = self.compare_headers(header_1, header_2)
        for header in headers:
            self.get_column(sheet, headers[header], row = 2)

    # might be better to put in Base
    def get_values(self, sheet, headers) -> list:
        '''
        get all cell value in a column after the header
        :param sheet: sheet to get values from
        :param column: column to get values from
        return list of all values in specified column
        '''
        column = []

        for col in sheet.iter_cols(values_only = True):
            for header in headers:
                if col == header:
                    colum.append(self.get_column(sheet, col, 2))
                else:
                    column.append('nothing')

        return column

    def combine_header_values(self, sheet, column, value_list) -> list:
        '''
        creates a list of headers as the amount of cells in that column
        combines that list with the list values
        '''
        header = self.get_column(sheet, column, row = 1)[0]
        headers = self.create_heading_list(header, len(value_list))

        return list(self.insert_every_n(headers, value_list, k = 1))

    def final_data(self, column, attributes, where_to_insert):
        return list(Base.insert_every_n(self, column, attributes, where_to_insert))
